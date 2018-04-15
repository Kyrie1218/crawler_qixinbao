#!/usr/bin/python3

# -*- coding:utf-8 -*-

import xlrd, arrow, urllib, re, os, requests

#from urllib.request import urlretrieve     #2018-02-25 SSL Error

import urllib3

from urllib.parse import quote

from selenium import webdriver

from bs4 import BeautifulSoup

from selenium.webdriver.common.desired_capabilities import DesiredCapabilities

from selenium.webdriver.common.action_chains import ActionChains

from selenium.webdriver.support.ui import WebDriverWait

from selenium.webdriver.support import expected_conditions as EC

from selenium.webdriver.common.by import By

from openpyxl.workbook import Workbook

import random, time

from PIL import Image, ImageFont, ImageDraw

import pytesseract

import os

import database



from requests.packages.urllib3.exceptions import InsecureRequestWarning



requests.packages.urllib3.disable_warnings(InsecureRequestWarning)



def openexcel(file):

    """

    open excel file

    :param file: excel file

    :return: excelojb

    """

    try:

        book = xlrd.open_workbook(file)

        return book

    except Exception as e:

        print("open excel file failed", str(e))





def readsheets(file):

    """

    read sheet

    :param file: excel obj

    :return: sheet obj

    """

    try:

        book = openexcel(file)

        sheet = book.sheets()

        return sheet

    except Exception as e:

        print("read sheet failed", str(e))





def readdata(sheet, n=0):

    """

    data read

    :param sheet: excel sheet

    :param n: rows

    :return: data list

    """

    dataset = []

    for r in range(sheet.nrows):

        col = sheet.cell(r, n).value

        # 如果有表头

        if r != 0:

            dataset.append(col)

    return dataset





def browserdriver():

    """

    start driver

    :return: driver obj

    """

    ua = "Mozilla/5.0 (Linux; Android 4.4.2; Nexus 4 Build/KOT49H) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/34.0.1847.114 Mobile Safari/537.36"

    profile = webdriver.FirefoxProfile()

    profile.set_preference(

        "general.useragent.override", ua)

    driver = webdriver.Firefox(profile, executable_path='lib/geckodriver.exe')

    return driver





# def browserdriver():

    """

    start driver

    :return: driver obj

    """

    """

    dcap = DesiredCapabilities.PHANTOMJS.copy()

    dcap['phantomjs.page.customHeaders.Referer'] = 'https://www.baidu.com/'

    dcap["phantomjs.page.settings.userAgent"] = 'Mozilla/5.0 (Linux; Android 4.4.2; Nexus 4 Build/KOT49H) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/34.0.1847.114 Mobile Safari/537.36'

    dcap["phantomjs.page.settings.loadImages"] = False

    dcap["phantomjs.page.customHeaders.User-Agent"] = 'Mozilla/5.0 (Linux; Android 4.4.2; Nexus 4 Build/KOT49H) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/34.0.1847.114 Mobile Safari/537.36'

    dcap["phantomjs.page.settings.disk-cache"] = True

    driver = webdriver.PhantomJS(executable_path='lib/phantomjs', desired_capabilities=dcap)

    return driver

"""





def request(url):

    header = {

        "Referer": "https://www.baidu.com/",

        "User-Agent": "Mozilla/5.0 (Linux; Android 4.4.2; Nexus 4 Build/KOT49H) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/34.0.1847.114 Mobile Safari/537.36"

    }

    fail = 1

    while fail < 31:

        try:

            r = requests.get(url, headers=header, timeout=10, verify=False)

            break

        except:

            fail += 1

            time.sleep(5)

    return r.text





def tyc_data(driver, url, keyword):

    """

    get Tianyancha Data

    :param driver: brower

    :param url: url

    :param keyword: keyword

    :return: tyc date

    """

    fails = 1

    while fails < 31:

        try:

            driver.set_page_load_timeout(10)

            driver.get("about:blank")

            driver.get(url)

            break

        except Exception as e:

            print(e)

            print('error wait 30 secs tyc_data1')

            time.sleep(30)

    else:

        raise

    try:

        WebDriverWait(driver, 60).until(

        EC.presence_of_element_located((By.CLASS_NAME, "footerV2"))

        )

    except Exception as e:

        print(e)

    finally:

        source = driver.page_source.encode("utf-8")

        tycsoup = BeautifulSoup(source, 'html.parser')

        name = tycsoup.select(
            "div > div > div#wap_header_top > div > div > div.f18")[0].get_text()


        cmname = name if len(name) > 0 else None

        print(cmname)
        return

        if cmname:

            company_url = "https://m.tianyancha.com" + tycsoup.select('div > div > div > div > a.query_name')[0].get('href')

            print(company_url)

            fails = 1

            while fails < 31:

                try:

                    driver.set_page_load_timeout(10)

                    driver.get("about:blank")

                    driver.get(company_url)

                    break

                except Exception as e:

                    print(e)

                    print('error wait 30 secs tyc_data2')

                    time.sleep(30)

            else:

                raise

            try:

                WebDriverWait(driver, 60).until(

                    EC.presence_of_element_located((By.CLASS_NAME, "footerV2"))

                )

            except Exception as e:

                print(e)

            finally:

                noinfo = ""

                source = driver.page_source.encode("utf-8")

                tycdata = BeautifulSoup(source, 'html.parser')

                binfo = []

                reginfo = tycdata.select("div.item-line > span")

                bsocplist = tycdata.select("div.item-line > span > span > span.hidden > div")

                gdinfo = ''

                dwinfo = ''

                gdlist = tycdata.select(

                    "div > div#_container_holder > div > div.content-container > div > div > a.in-block"

                )

                if(len(gdlist) != 0):

                    for gdname in gdlist:

                        gdinfo += gdname.get_text() + "|,|"

                dw_list = tycdata.select(

                    "div.content-container > div > a > span.text-click-color"

                )

                if(len(dw_list) != 0):

                    for dw in dw_list:

                        dwinfo += dw.get_text() + "|,|"

                bscop = bsocplist[0].get_text() if len(bsocplist) > 0 else None

                if (reginfo[11].get_text()) == "民办非企业单位" or (reginfo[11].get_text()) == "社会团体":

                    binfo = [

                        cmname,

                        noinfo,

                        reginfo[1].get_text() if len(reginfo[1].get_text()) > 0 else None,

                        # 修改填充项

                        # regdecode(maping, reginfo[3].get_text()) if len(reginfo[3].get_text()) > 0 else None,

                        # regdecode(maping, reginfo[5].get_text()) if len(reginfo[5].get_text()) > 0 else None,

                        reginfo[5].get_text() if len(reginfo[5].get_text()) > 0 else None,

                        reginfo[3].get_text() if len(reginfo[3].get_text()) > 0 else None,

                        noinfo,

                        reginfo[7].get_text(),

                        noinfo,

                        reginfo[9].get_text(),

                        reginfo[11].get_text(),

                        noinfo,

                        noinfo,

                        noinfo,

                        reginfo[13].get_text(),

                        noinfo,

                        noinfo

                    ]

                else:

                    binfo = [

                        cmname,

                        reginfo[3].get_text() if len(reginfo[3].get_text()) > 0 else None,

                        reginfo[1].get_text() if len(reginfo[1].get_text()) > 0 else None,

                        #regdecode(maping, reginfo[7].get_text()) if len(reginfo[7].get_text()) > 0 else None,

                        #regdecode(maping, reginfo[5].get_text()) if len(reginfo[5].get_text()) > 0 else None,

                        #regdecode(maping, reginfo[23].get_text()) if len(reginfo[23].get_text()) > 0 else None,

                        reginfo[13].get_text() if len(reginfo[13].get_text()) > 0 else None,

                        reginfo[15].get_text() if len(reginfo[15].get_text()) > 0 else None,

                        reginfo[17].get_text() if len(reginfo[17].get_text()) > 0 else None,

                        reginfo[11].get_text() if len(reginfo[11].get_text()) > 0 else None,

                        reginfo[19].get_text() if len(reginfo[19].get_text()) > 0 else None,

                        reginfo[9].get_text() if len(reginfo[9].get_text()) > 0 else None,

                        reginfo[21].get_text() if len(reginfo[21].get_text()) > 0 else None,

                        reginfo[25].get_text() if len(reginfo[25].get_text()) > 0 else None,

                        reginfo[27].get_text() if len(reginfo[27].get_text()) > 0 else None,

                        bscop

                    ]

                if(len(gdinfo) != 0):

                    binfo.append(gdinfo)

                else:

                    binfo.append('暂无')

                    print("暂无股东信息")

                if(len(dw_list) != 0):

                    binfo.append(dwinfo)

                else:

                    binfo.append('暂无')

                    print("暂无对外投资信息")

                for x in binfo:

                    print(x)

                return binfo

        else:

            print('暂无信息')

            binfo = [keyword, "暂无信息"]

            return binfo

def company_list(driver, url, keyword):

    fails = 1

    while fails < 31:

        try:

            driver.set_page_load_timeout(10)

            driver.get("about:blank")

            driver.get(url)

            break

        except Exception as e:

            print(e)

            print('error wait 30 secs tyc_data1')

            time.sleep(30)

    else:

        raise

    try:

        WebDriverWait(driver, 60).until(

        EC.presence_of_element_located((By.CLASS_NAME, "footerV2"))

        )

    except Exception as e:

        print(e)

    finally:

        source = driver.page_source.encode("utf-8")

        tycsoup = BeautifulSoup(source, 'html.parser')

        company = tycsoup.select(
            "div > div.search_result_container > div.new-border-bottom > div > div > a")

        data = []

        for link in company:
            data.append(link.get('href'))

    return data


# 入口函数
def main2(logfile, excelfile):

    try:
        driver = browserdriver()
    except Exception as e:
        print(e)

    # 获取当前时间
    now = arrow.now()

    # 导出excel文件名称
    newexcelfile = "" + arrow.now().format("YYYY-MM-DD HH_mm_ss") + ".xlsx"

    # 新建一个excel操作类
    wb = Workbook()

    ws = wb.active

    ws.append([
        "公司名称", "公司状态", "法人名称", "注册资本", "注册时间", "核准时间", "工商注册号", "组织机构代码", "信用识别代码",
        "公司类型", "纳税人识别号", "行业", "营业期限", "登记机关", "注册地址", "经营范围", "股东", "对外投资"
    ])

    dcount = 0

    for sheet in readsheets('company_ids.xlsx'):

        for cmyname in readdata(sheet):

            keyword = quote(cmyname.encode("utf-8"))

            tycurl = "https://m.tianyancha.com" + keyword

            binfo = tyc_data(driver, tycurl, cmyname)
            print(binfo)
            return
            if binfo is None:

                print("Error", binfo, "is None")

                pass

            else:

                ws.append(binfo)

                wb.save(filename=newexcelfile)

            a = random.randint(0, 0)

            print('采集完毕，等待' + str(a) + '秒')

            time.sleep(a)

            dcount += 1

            if dcount == 150:

                try:

                    driver.quit()

                except Exception as e:

                    print(e)

                finally:

                    dcount = 0

                    driver = browserdriver()

            else:

                pass

    wb.save(filename=newexcelfile)

    driver.quit()


# 入口函数
def main(logfile, excelfile):

    try:
        driver = browserdriver()
    except Exception as e:
        print(e)

    # 获取当前时间
    now = arrow.now()

    # 导出excel文件名称
    newexcelfile = "company_ids.xlsx"

    # 创建一个工作表
    wb = Workbook()

    # 找到活动的sheet页。空的excel表默认的sheet页就叫Sheet，如果想改名字，可以直接给title属性赋值。
    ws = wb.active

    dcount = 0

    for sheet in readsheets('company.xlsx'):

        for cmyname in readdata(sheet):

            keyword = quote(cmyname.encode("utf-8"))

            tycurl = "https://m.tianyancha.com/search?key=" + keyword + "&checkFrom=searchBox"

            company_ids = company_list(driver, tycurl, cmyname)

            if company_ids is None:
                print("Error", company_ids, "is None")
                pass
            else:
                ws.append(company_ids)
                wb.save(filename=newexcelfile)

            a = random.randint(0, 0)

            print('采集完毕，等待' + str(a) + '秒')

            time.sleep(a)

            dcount += 1

            if dcount == 150:

                try:

                    driver.quit()

                except Exception as e:

                    print(e)

                finally:

                    dcount = 0

                    driver = browserdriver()

            else:

                pass

    wb.save(filename=newexcelfile)

    driver.quit()


if __name__ == '__main__':

    db = Databse
    print(db)

    # logfile = 'log.txt'

    # excel = 'company.xlsx'
    # excel2 = 'company_ids.xlsx'

    # # main(logfile, excel)

    # main2(logfile, excel2)
